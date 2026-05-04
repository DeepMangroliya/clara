"""Scoreboard XLSX → JSON converter.

Reads the clinic performance scoreboard workbook and emits a structured JSON
document that preserves every cell value alongside its semantic metadata
(focus, source, role, target, section). The output is designed to be
queryable rather than a verbatim dump: metrics carry stable keys, weekly
data is keyed by ISO date, and the schema is versioned for forward
compatibility.

Public API
----------
    convert_workbook(input_path, output_path, *, sheet_name=None) -> ScoreboardDocument

Command line
------------
    python convert.py INPUT.xlsx [-o OUTPUT.json] [--sheet NAME] [-v]

Exit codes
----------
    0  Success.
    2  Input file missing or unreadable.
    3  Workbook is structurally invalid (no sheets, no header rows, etc.).
    4  Unexpected internal error.

Module attributes
-----------------
    SCHEMA_VERSION: Semantic version of the emitted JSON document. Bump on
        any breaking change to the output shape.
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import tempfile
import zipfile
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterator, Sequence

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Module constants
# ---------------------------------------------------------------------------

#: Semantic version of the emitted JSON document. Bump the major component on
#: any change that an existing consumer cannot tolerate (renamed/removed
#: fields, changed types). Bump the minor component on additive changes.
SCHEMA_VERSION = "1.0.0"

#: Row layout of the scoreboard. Rows are 1-indexed to match Excel's own
#: addressing, which keeps debugging against the source file straightforward.
#: Row 6 holds the literal word "Target" and carries no per-metric data, so
#: it is intentionally omitted here.
ROW_BANNER = 1   # Section banners (e.g. the merged "PHONE PERFORMANCE" cell).
ROW_METRIC = 2   # Human-readable metric name; the canonical column header.
ROW_FOCUS = 3    # Focus area: Financial, Marketing, Caseload, etc.
ROW_SOURCE = 4   # Upstream data source: EMR, Jane, CallHero, Formula, etc.
ROW_ROLE = 5     # Person responsible for the metric.
ROW_TARGET = 7   # Target value or descriptive note.
FIRST_DATA_ROW = 8

#: Column A holds the week-ending date for each data row.
DATE_COLUMN = 1

#: Discipline prefixes used to recognise per-service blocks (PT, RMT, etc.).
#: Order matters: longer/more specific prefixes are checked first to avoid
#: matching a shorter prefix against a longer service name.
DISCIPLINE_PREFIXES: tuple[tuple[str, str], ...] = (
    ("Pelvic Health", "Pelvic Health"),
    ("CHIRO ", "CHIRO"),
    ("Chiro ", "CHIRO"),
    ("RMT ", "RMT"),
    ("PT ", "PT"),
)

#: Default section name applied when no banner or discipline prefix is found.
DEFAULT_SECTION = "Clinic Wide"

logger = logging.getLogger("scoreboard.convert")


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class Metric:
    """Definition of a single scoreboard column.

    Attributes:
        key: Stable, JSON-safe identifier matching ``[a-z0-9_]+``. Used as
            the join key between :class:`Metric` and :class:`WeekRecord`.
        label: Original human-readable header from row 2, with whitespace
            normalised but otherwise verbatim. ``None`` when the source
            column has data but no header.
        section: Logical grouping the metric belongs to (e.g. ``"PT"``).
        focus: Functional area from row 3 (``"Financial"``, ``"Marketing"`` …).
        source: Upstream system from row 4 (``"EMR"``, ``"Jane"`` …).
        role: Person responsible from row 5.
        target: Target value or descriptive note from row 7.
        column: Excel column letter (``"B"``, ``"AJ"`` …) for traceability
            back to the source workbook.
        description: Long-form definition of the metric, captured from the
            cell comment on the row-2 header. ``None`` when no comment is
            present. These comments are how the source workbook documents
            what each KPI actually means and are valuable context for any
            downstream consumer.
        url: External URL captured from a ``=HYPERLINK(...)`` formula in
            any of rows 1–7. The display text of such a hyperlink is
            already preserved on whichever named field it lives in
            (``source``, ``target`` …); this field surfaces the underlying
            link so it isn't silently dropped. ``None`` when no hyperlink
            applies to this column.
        hidden: ``True`` if the source column is hidden in Excel. The
            column's data is still preserved end-to-end; this flag lets a
            downstream dashboard or report mirror the source's visibility
            choices when desired.
    """

    key: str
    label: str | None
    section: str
    focus: str | None
    source: str | None
    role: str | None
    target: Any
    column: str
    description: str | None = None
    url: str | None = None
    hidden: bool = False


@dataclass(frozen=True)
class Section:
    """A named group of contiguous metric columns."""

    name: str
    metric_keys: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class WeekRecord:
    """One row of weekly data, keyed by metric key.

    ``week_ending`` is an ISO-8601 date string. ``values`` only contains
    keys for cells that actually held data, so absence of a key means
    "no data recorded" rather than "value is null".
    """

    week_ending: str
    values: dict[str, Any]


@dataclass(frozen=True)
class ScoreboardDocument:
    """Top-level conversion result. Serialises to the public JSON schema."""

    source_file: str
    sheet: str
    schema_version: str
    generated_at: str
    sections: list[Section]
    metrics: list[Metric]
    weeks: list[WeekRecord]

    def to_json_dict(self) -> dict[str, Any]:
        """Return a plain ``dict`` suitable for ``json.dumps``."""
        return asdict(self)


# ---------------------------------------------------------------------------
# Custom exceptions
# ---------------------------------------------------------------------------


class ScoreboardError(Exception):
    """Base class for all converter-specific errors."""


class WorkbookStructureError(ScoreboardError):
    """The workbook does not match the expected scoreboard layout."""


# ---------------------------------------------------------------------------
# Cell-level helpers
# ---------------------------------------------------------------------------


# Pre-compiled patterns. Compiling once at import is cheap and avoids
# re-compilation in inner loops over hundreds of cells.
_RE_INT = re.compile(r"-?\d+")
_RE_FLOAT = re.compile(r"-?\d+\.\d+")
_RE_NON_KEY_CHAR = re.compile(r"[^a-z0-9_]")
_RE_MULTI_UNDERSCORE = re.compile(r"_+")
_RE_EXCEL_ERROR = re.compile(r"^#[A-Z?/]+!$")
_RE_WHITESPACE = re.compile(r"\s+")


def _slugify(text: str) -> str:
    """Convert a header label into a stable ``[a-z0-9_]+`` identifier.

    A small set of meaningful symbols is mapped to readable words before
    stripping (``%`` → ``pct``, ``#`` → ``num``, ``&`` → ``and``,
    ``+`` → ``plus``, ``/`` → ``per``) so that ``"NAR % Collected at Ax"``
    and ``"NAR Collected at Ax"`` produce distinct, self-describing keys
    rather than colliding.

    The original label is always preserved verbatim on the :class:`Metric`,
    so this function's output is purely for programmatic use.
    """
    if not text:
        return "unnamed"

    normalised = text.replace("\n", " ").replace("\\", " ")
    for symbol, replacement in (
        ("%", " pct "),
        ("#", " num "),
        ("&", " and "),
        ("+", " plus "),
        ("/", " per "),
    ):
        normalised = normalised.replace(symbol, replacement)

    # Lowercase, replace whitespace runs with underscores, drop anything
    # outside the identifier alphabet, and collapse repeated underscores.
    normalised = _RE_WHITESPACE.sub("_", normalised.strip()).lower()
    normalised = _RE_NON_KEY_CHAR.sub("", normalised)
    normalised = _RE_MULTI_UNDERSCORE.sub("_", normalised).strip("_")
    return normalised or "unnamed"


def _normalize_cell(value: Any) -> Any:
    """Convert one openpyxl cell value into a JSON-safe Python value.

    Rules:
      * ``None`` and empty/whitespace strings → ``None``.
      * Excel error tokens (``"#REF!"``, ``"#DIV/0!"`` …) →
        ``{"error": "<token>"}`` so consumers can distinguish "broken
        formula" from "no data".
      * Numeric strings (``"2.87"``, ``"42"``) are coerced to ``float`` /
        ``int``. The source workbook mixes string- and number-typed
        numerics for the same metric across rows; leaving them mixed would
        force every consumer to re-detect the type.
      * Dates and datetimes → ISO-8601 strings.
      * Booleans pass through as ``bool`` (checked before ``int`` because
        ``bool`` is an ``int`` subclass in Python).
      * All other values pass through unchanged, with a debug log for any
        types we did not anticipate.
    """
    if value is None:
        return None

    if isinstance(value, bool):
        return value

    if isinstance(value, datetime):
        # If there is no time-of-day component, prefer a date-only string;
        # this matches how the workbook displays week-ending values.
        if value.time() == datetime.min.time():
            return value.date().isoformat()
        return value.isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, (int, float)):
        return value

    if isinstance(value, str):
        stripped = _RE_WHITESPACE.sub(" ", value.replace("\n", " ")).strip()
        if not stripped or stripped == "\\":
            return None
        if _RE_EXCEL_ERROR.match(stripped):
            return {"error": stripped}
        if _RE_FLOAT.fullmatch(stripped):
            return float(stripped)
        if _RE_INT.fullmatch(stripped):
            return int(stripped)
        return stripped

    logger.debug("Unhandled cell type %s; coercing via str()", type(value).__name__)
    return str(value)


# Pattern to capture the URL embedded in an ``=HYPERLINK("url", "label")``
# formula. Captures only the URL; the visible label survives as the cell's
# cached display value and is read separately via the data-only workbook.
_RE_HYPERLINK = re.compile(
    r'^=HYPERLINK\(\s*"([^"]+)"', re.IGNORECASE
)

# Timestamp line in cell comments — matches strings like ``(2026-04-29 16:56:49)``.
_RE_COMMENT_TIMESTAMP = re.compile(r"\(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\)")


def _extract_comment(cell: Any) -> str | None:
    """Return a cell's comment text with whitespace normalised.

    Excel comments in this workbook follow a consistent shape: a header
    line (``======``), an internal ID, a timestamp, then the actual
    descriptive text. The header noise is dropped so the field carries
    only the human-readable definition.
    """
    if cell is None or cell.comment is None:
        return None

    raw = cell.comment.text or ""
    if not raw.strip():
        return None

    # Strip the boilerplate prelude ('======', 'ID#...', timestamp) that
    # every comment in this workbook carries. Anything still containing
    # useful text after that filter is the real description.
    cleaned_lines: list[str] = []
    for line in raw.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith("======"):
            continue
        if stripped.startswith("ID#"):
            continue
        if _RE_COMMENT_TIMESTAMP.fullmatch(stripped):
            continue
        cleaned_lines.append(stripped)

    if not cleaned_lines:
        return None
    return _RE_WHITESPACE.sub(" ", " ".join(cleaned_lines))


def _extract_hyperlink_url(formula_cell: Any) -> str | None:
    """Pull the URL out of a ``=HYPERLINK("url", "text")`` formula.

    ``formula_cell`` must come from a workbook loaded with
    ``data_only=False``; otherwise its value is the formula's cached
    result rather than the formula itself.
    """
    if formula_cell is None:
        return None
    value = formula_cell.value
    if not isinstance(value, str):
        return None
    match = _RE_HYPERLINK.match(value)
    return match.group(1) if match else None


# ---------------------------------------------------------------------------
# Workbook structure helpers
# ---------------------------------------------------------------------------


def _effective_max_row(ws: Worksheet) -> int:
    """Return the index of the last row containing any non-empty cell.

    ``Worksheet.max_row`` can be inflated by stale formatting on otherwise
    empty rows. Iterating from the bottom up and stopping at the first row
    with content gives a stable, content-driven bound for downstream loops.
    """
    for row_idx in range(ws.max_row, 0, -1):
        for cell in ws[row_idx]:
            if cell.value is None or cell.value == "":
                continue
            if isinstance(cell.value, str) and not cell.value.strip():
                continue
            return row_idx
    return 0


def _column_is_empty(ws: Worksheet, col_idx: int, last_row: int) -> bool:
    """Return ``True`` iff column ``col_idx`` has no content in rows 1..last_row."""
    for row_idx in range(1, last_row + 1):
        value = ws.cell(row=row_idx, column=col_idx).value
        if value is None or value == "":
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return False
    return True


def _resolve_section_name(ws: Worksheet, run: Sequence[int]) -> str:
    """Choose a section name for a contiguous run of columns.

    Rule precedence:

      1. A non-empty banner cell in row 1 anywhere in the run.
      2. A discipline prefix on the first metric label of the run.
      3. Fall back to :data:`DEFAULT_SECTION`.

    Inferring section names from "first word of the first metric" is
    deliberately not attempted: it produces noise like ``"FB"`` or
    ``"Cancelled"`` for groups that are not really their own thing.
    """
    for col_idx in run:
        banner = ws.cell(row=ROW_BANNER, column=col_idx).value
        if banner is not None and str(banner).strip():
            text = str(banner).strip()
            return text.title() if text.isupper() else text

    first_label = ws.cell(row=ROW_METRIC, column=run[0]).value
    if isinstance(first_label, str):
        cleaned = first_label.strip().replace("\n", " ")
        for prefix, name in DISCIPLINE_PREFIXES:
            if cleaned.startswith(prefix):
                return name

    return DEFAULT_SECTION


def _iter_column_runs(
    ws: Worksheet, last_row: int, last_col: int
) -> Iterator[list[int]]:
    """Yield contiguous runs of non-empty columns, splitting on spacers.

    Column A (the date column) is excluded; the caller handles dates
    separately.
    """
    current: list[int] = []
    for col_idx in range(DATE_COLUMN + 1, last_col + 1):
        if _column_is_empty(ws, col_idx, last_row):
            if current:
                yield current
                current = []
            continue
        current.append(col_idx)
    if current:
        yield current


# ---------------------------------------------------------------------------
# Extraction
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class _ExtractionContext:
    """Bundles the worksheets and metadata needed to build a :class:`Metric`.

    ``data_ws`` is the data-only view (cached formula values); ``formula_ws``
    is the same sheet loaded with formula text intact. Both are required
    because the description and URL fields live in different parts of the
    workbook than the cached cell values.
    """

    data_ws: Worksheet
    formula_ws: Worksheet
    hidden_columns: frozenset[str]


def _extract_metric(
    ctx: _ExtractionContext, col_idx: int, section_name: str
) -> Metric:
    """Build a :class:`Metric` from the header rows above ``col_idx``."""
    ws = ctx.data_ws
    label = _normalize_cell(ws.cell(row=ROW_METRIC, column=col_idx).value)
    column_letter = get_column_letter(col_idx)

    # Some columns have data but no header label (e.g. column U in the
    # reference workbook). Synthesise a positional key so the data is not
    # silently dropped, and leave ``label`` as ``None`` to signal absence.
    label_for_key = label if isinstance(label, str) and label else f"col_{column_letter}"

    # Cell comments on the row-2 header carry the canonical definition of
    # what each metric measures (e.g. "Enter the total revenue billed
    # across all services for a given week"). We surface those as
    # ``description``.
    description = _extract_comment(ws.cell(row=ROW_METRIC, column=col_idx))

    # The source workbook embeds Google Sheets links via
    # ``=HYPERLINK("url", "label")`` formulas in the source row. The
    # cached value is just the visible label; the URL itself lives in the
    # formula text and is only readable from the formula-mode worksheet.
    url: str | None = None
    for header_row in (ROW_BANNER, ROW_METRIC, ROW_FOCUS, ROW_SOURCE,
                       ROW_ROLE, ROW_TARGET):
        url = _extract_hyperlink_url(
            ctx.formula_ws.cell(row=header_row, column=col_idx)
        )
        if url:
            break

    return Metric(
        key=_slugify(label_for_key),
        label=label if isinstance(label, str) else None,
        section=section_name,
        focus=_normalize_cell(ws.cell(row=ROW_FOCUS, column=col_idx).value),
        source=_normalize_cell(ws.cell(row=ROW_SOURCE, column=col_idx).value),
        role=_normalize_cell(ws.cell(row=ROW_ROLE, column=col_idx).value),
        target=_normalize_cell(ws.cell(row=ROW_TARGET, column=col_idx).value),
        column=column_letter,
        description=description,
        url=url,
        hidden=column_letter in ctx.hidden_columns,
    )


def _replace_metric_key(metric: Metric, new_key: str) -> Metric:
    """Return a copy of ``metric`` with ``key`` replaced.

    Required because :class:`Metric` is frozen; we cannot mutate ``key``
    in place. ``dataclasses.replace`` would also work but explicit
    construction documents intent at the call site.
    """
    return Metric(
        key=new_key,
        label=metric.label,
        section=metric.section,
        focus=metric.focus,
        source=metric.source,
        role=metric.role,
        target=metric.target,
        column=metric.column,
        description=metric.description,
        url=metric.url,
        hidden=metric.hidden,
    )


def _disambiguate_keys(metrics: list[Metric]) -> list[Metric]:
    """Return a new list of metrics with collision-free keys.

    The same label may appear multiple times in the workbook:

      * Across discipline blocks (``"PVA (4 wk avg)"`` recurs under PT,
        RMT, Chiro, and Pelvic Health).
      * Within a single section (``"Utilization"`` appears twice in the
        Clinic Wide run).

    Strategy: section-prefix first, then break any remaining ties with
    the Excel column letter. The column letter is unique by construction,
    so the second pass is bounded and terminates.
    """
    counts: dict[str, int] = {}
    for metric in metrics:
        counts[metric.key] = counts.get(metric.key, 0) + 1

    first_pass: list[Metric] = []
    for metric in metrics:
        if counts[metric.key] > 1:
            new_key = f"{_slugify(metric.section)}__{metric.key}"
            first_pass.append(_replace_metric_key(metric, new_key))
        else:
            first_pass.append(metric)

    counts.clear()
    for metric in first_pass:
        counts[metric.key] = counts.get(metric.key, 0) + 1

    result: list[Metric] = []
    for metric in first_pass:
        if counts[metric.key] > 1:
            new_key = f"{metric.key}_{metric.column.lower()}"
            result.append(_replace_metric_key(metric, new_key))
        else:
            result.append(metric)
    return result


def _extract_weeks(
    ws: Worksheet,
    metrics: Sequence[Metric],
    last_row: int,
) -> list[WeekRecord]:
    """Produce one :class:`WeekRecord` per data row.

    Cells whose normalised value is ``None`` are omitted from
    :attr:`WeekRecord.values` to keep the output compact and to make
    ``key in record.values`` a meaningful predicate.
    """
    # Pre-compute (key, col_idx) pairs once so the inner per-row loop is
    # tight; avoids re-converting column letters on every cell access.
    key_columns: list[tuple[str, int]] = [
        (m.key, column_index_from_string(m.column)) for m in metrics
    ]

    weeks: list[WeekRecord] = []
    for row_idx in range(FIRST_DATA_ROW, last_row + 1):
        date_cell = ws.cell(row=row_idx, column=DATE_COLUMN).value
        if date_cell in (None, ""):
            continue

        if isinstance(date_cell, datetime):
            week_ending = date_cell.date().isoformat()
        elif isinstance(date_cell, date):
            week_ending = date_cell.isoformat()
        else:
            normalised = _normalize_cell(date_cell)
            if normalised is None:
                logger.warning(
                    "Row %d has a non-empty date cell that normalised to None; skipping.",
                    row_idx,
                )
                continue
            week_ending = str(normalised)

        values: dict[str, Any] = {}
        for key, col_idx in key_columns:
            cell_value = _normalize_cell(ws.cell(row=row_idx, column=col_idx).value)
            if cell_value is not None:
                values[key] = cell_value

        weeks.append(WeekRecord(week_ending=week_ending, values=values))

    return weeks


def _build_sections(metrics: Sequence[Metric]) -> list[Section]:
    """Group metrics by section, preserving first-occurrence order.

    Multiple non-adjacent column runs may map to the same logical section
    (notably ``"Clinic Wide"``); they are collapsed into a single
    :class:`Section` whose ``metric_keys`` lists the union in sheet order.
    """
    order: list[str] = []
    by_section: dict[str, list[str]] = {}
    for metric in metrics:
        if metric.section not in by_section:
            order.append(metric.section)
            by_section[metric.section] = []
        by_section[metric.section].append(metric.key)
    return [Section(name=name, metric_keys=by_section[name]) for name in order]


# ---------------------------------------------------------------------------
# Top-level conversion
# ---------------------------------------------------------------------------


def _select_sheet(workbook: Any, requested: str | None) -> Worksheet:
    """Return the worksheet to convert.

    Validates that the workbook has at least one sheet and that
    ``requested`` (if given) actually exists.
    """
    if not workbook.sheetnames:
        raise WorkbookStructureError("Workbook contains no sheets.")
    if requested is None:
        return workbook[workbook.sheetnames[0]]
    if requested not in workbook.sheetnames:
        raise WorkbookStructureError(
            f"Sheet {requested!r} not found. Available: {workbook.sheetnames!r}"
        )
    return workbook[requested]


def _validate_layout(ws: Worksheet, last_row: int) -> None:
    """Sanity-check that the worksheet has the expected scoreboard shape."""
    if ws.max_column < DATE_COLUMN + 1:
        raise WorkbookStructureError(
            f"Sheet {ws.title!r} has no metric columns "
            f"(max column = {ws.max_column})."
        )
    if last_row < FIRST_DATA_ROW:
        raise WorkbookStructureError(
            f"Sheet {ws.title!r} has no data rows "
            f"(effective max row = {last_row}, expected at least {FIRST_DATA_ROW})."
        )
    if not any(
        ws.cell(row=ROW_METRIC, column=col_idx).value
        for col_idx in range(DATE_COLUMN + 1, ws.max_column + 1)
    ):
        raise WorkbookStructureError(
            f"Sheet {ws.title!r} has no metric headers in row {ROW_METRIC}."
        )


def convert_workbook(
    input_path: Path,
    output_path: Path | None = None,
    *,
    sheet_name: str | None = None,
) -> ScoreboardDocument:
    """Convert a scoreboard workbook to the structured JSON document.

    Args:
        input_path: Path to the source ``.xlsx`` file.
        output_path: Where to write the JSON. If ``None``, no file is
            written; the caller still receives the in-memory document.
        sheet_name: Specific sheet to convert. Defaults to the first
            sheet in the workbook.

    Returns:
        The fully-populated :class:`ScoreboardDocument`.

    Raises:
        FileNotFoundError: ``input_path`` does not exist or is not a file.
        PermissionError: ``input_path`` is not readable, or
            ``output_path`` is not writable.
        WorkbookStructureError: The workbook does not match the expected
            scoreboard layout.
    """
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    if not input_path.is_file():
        raise FileNotFoundError(f"Input path is not a file: {input_path}")

    logger.info("Loading workbook: %s", input_path)
    # The converter needs two views of the workbook:
    #
    #   * ``data_only=True``   gives cached formula results (the numbers a
    #     human sees opening the file in Excel) — used for all cell values.
    #   * ``data_only=False``  gives raw formula text — used to recover the
    #     URL out of ``=HYPERLINK("...")`` cells and any other metadata
    #     that is encoded as a formula rather than a value.
    #
    # The scoreboard is small enough that loading it twice is essentially
    # free. ``read_only=True`` would be faster for very large files but
    # disables features we rely on (cell comments and column dimensions).
    #
    # ``load_workbook`` raises a wide variety of exceptions for malformed
    # input: ``zipfile.BadZipFile`` for non-zip data, ``KeyError`` for
    # missing internal parts, ``ValueError`` for unsupported features,
    # and ``OSError`` for I/O problems. We translate all of them into
    # ``WorkbookStructureError`` so the CLI can surface a clean message
    # instead of an internal traceback. ``PermissionError`` (an
    # ``OSError`` subclass) is allowed to propagate so the caller can
    # distinguish "cannot read" from "wrong format".
    try:
        workbook = load_workbook(filename=input_path, data_only=True)
        formula_workbook = load_workbook(filename=input_path, data_only=False)
    except PermissionError:
        raise
    except (OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
        raise WorkbookStructureError(
            f"Failed to open {input_path.name}: {exc}"
        ) from exc

    worksheet = _select_sheet(workbook, sheet_name)
    formula_worksheet = _select_sheet(formula_workbook, worksheet.title)
    last_row = _effective_max_row(worksheet)
    last_col = worksheet.max_column

    logger.debug(
        "Sheet %r: max_col=%d, effective_last_row=%d (raw max_row=%d)",
        worksheet.title, last_col, last_row, worksheet.max_row,
    )

    _validate_layout(worksheet, last_row)

    # Snapshot which columns are hidden in the source. ``column_dimensions``
    # is keyed by column letter; absence from the dict means "default,
    # i.e. visible". Computing this once avoids repeated dict lookups
    # during metric extraction.
    hidden_columns = frozenset(
        letter for letter, dim in worksheet.column_dimensions.items()
        if dim.hidden
    )
    logger.debug("Hidden columns: %s", sorted(hidden_columns))

    ctx = _ExtractionContext(
        data_ws=worksheet,
        formula_ws=formula_worksheet,
        hidden_columns=hidden_columns,
    )

    # Walk the columns left-to-right, splitting on spacer columns.
    metrics: list[Metric] = []
    for run in _iter_column_runs(worksheet, last_row, last_col):
        section_name = _resolve_section_name(worksheet, run)
        for col_idx in run:
            metrics.append(_extract_metric(ctx, col_idx, section_name))

    metrics = _disambiguate_keys(metrics)
    sections = _build_sections(metrics)
    weeks = _extract_weeks(worksheet, metrics, last_row)

    document = ScoreboardDocument(
        source_file=input_path.name,
        sheet=worksheet.title,
        schema_version=SCHEMA_VERSION,
        generated_at=_utc_now_iso(),
        sections=sections,
        metrics=metrics,
        weeks=weeks,
    )

    logger.info(
        "Converted: %d metrics across %d sections, %d weeks.",
        len(metrics), len(sections), len(weeks),
    )

    if output_path is not None:
        _write_json_atomic(output_path, document.to_json_dict())
        logger.info("Wrote %s (%d bytes)", output_path, output_path.stat().st_size)

    return document


def _utc_now_iso() -> str:
    """Current UTC time, ISO-8601 with a trailing ``Z`` and no microseconds."""
    return (
        datetime.now(timezone.utc)
        .replace(microsecond=0)
        .isoformat()
        .replace("+00:00", "Z")
    )


def _write_json_atomic(path: Path, payload: dict[str, Any]) -> None:
    """Serialise ``payload`` to ``path`` atomically.

    Writes to a sibling tempfile and renames into place, so a partial
    write cannot leave a half-written ``output.json`` for downstream
    consumers. The temp file is created on the same filesystem as the
    destination so ``os.replace`` is guaranteed atomic on POSIX.

    ``tempfile.mkstemp`` creates files with mode ``0600`` for security.
    We relax that to respect the process umask (typically ``0644``) so
    the output is readable by other users and tools — the expected
    behaviour for a generated artifact.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    encoded = json.dumps(payload, indent=2, ensure_ascii=False)

    tmp_fd, tmp_name = tempfile.mkstemp(
        prefix=path.name + ".", suffix=".tmp", dir=str(path.parent)
    )
    try:
        with os.fdopen(tmp_fd, "w", encoding="utf-8") as fh:
            fh.write(encoded)
            fh.write("\n")
            fh.flush()
            os.fsync(fh.fileno())
        # Apply user umask to a 0666 baseline so the result is typically
        # 0644. ``os.umask`` returns the previous value; we restore it
        # immediately to avoid side-effecting the rest of the process.
        umask = os.umask(0)
        os.umask(umask)
        os.chmod(tmp_name, 0o666 & ~umask)
        os.replace(tmp_name, path)
    except Exception:
        # Best-effort cleanup; the original file (if any) is untouched.
        try:
            os.unlink(tmp_name)
        except OSError:
            pass
        raise


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="convert.py",
        description="Convert the clinic scoreboard XLSX into structured JSON.",
    )
    parser.add_argument(
        "input",
        type=Path,
        help="Path to the source .xlsx workbook.",
    )
    parser.add_argument(
        "-o", "--output",
        type=Path,
        default=Path("output.json"),
        help="Where to write the JSON document (default: %(default)s).",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Sheet name to convert (default: the first sheet in the workbook).",
    )
    parser.add_argument(
        "-v", "--verbose",
        action="count",
        default=0,
        help="Increase log verbosity (-v for INFO, -vv for DEBUG).",
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s schema {SCHEMA_VERSION}",
    )
    return parser.parse_args(argv)


def _setup_logging(verbosity: int) -> None:
    if verbosity >= 2:
        level = logging.DEBUG
    elif verbosity == 1:
        level = logging.INFO
    else:
        level = logging.WARNING
    logging.basicConfig(
        level=level,
        format="%(asctime)s %(levelname)-7s %(name)s: %(message)s",
        datefmt="%H:%M:%S",
    )


def main(argv: Sequence[str] | None = None) -> int:
    """Command-line entry point. Returns a process exit code."""
    try:
        args = _parse_args(argv if argv is not None else sys.argv[1:])
    except SystemExit as exc:
        # argparse calls sys.exit on -h/--help/--version and on errors.
        return int(exc.code) if exc.code is not None else 0

    _setup_logging(args.verbose)

    try:
        document = convert_workbook(
            args.input, args.output, sheet_name=args.sheet
        )
    except FileNotFoundError as exc:
        logger.error("%s", exc)
        return 2
    except PermissionError as exc:
        logger.error("Permission denied: %s", exc)
        return 2
    except WorkbookStructureError as exc:
        logger.error("Invalid workbook: %s", exc)
        return 3
    except Exception:  # pragma: no cover - defensive last-resort guard
        logger.exception("Unexpected error during conversion.")
        return 4

    # Always print a one-line summary on stdout regardless of log level
    # so the script is friendly for shell pipelines.
    print(
        f"Wrote {args.output} "
        f"({len(document.metrics)} metrics, "
        f"{len(document.sections)} sections, "
        f"{len(document.weeks)} weeks)."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())